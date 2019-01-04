#!/usr/bin/env python

import getopt
import sys
import os
os.environ["DARTSENSE_ENV"] = "DEV"

import sqlite3

from pprint import pprint
from dumper import dump

import openpyxl
import dateutil.parser

sys.path.append(os.path.join(os.path.dirname(__file__), "../lib"))
import dartsense
import dartsense.player
import dartsense.event
import dartsense.match
import dartsense.competition


players = {}


def find_player(player_name, competition_id):
    speler = None
    if player_name in players:
        speler = players[player_name]
    else:
        player_list = dartsense.player.PlayerList(
            filters={'competition': competition_id},
            search=player_name
        )

        # dump(player_list)

        if len(player_list) == 1:
            speler = players[player_name] = player_list[0]
        elif(len(player_list) > 1):
            for player in player_list:
                # pprint('0',player.name)
                # pprint('1',player_name)
                if player.name == player_name:
                    speler = player
                    break

    if not speler:
        # maak nieuwe speler aan
        speler = dartsense.player.Player(name=player_name)
        speler.save()

    return speler


def main(argv):

    dartsense.db.exec_sql('DELETE FROM event WHERE event_id > 0')
    dartsense.db.exec_sql('DELETE FROM match WHERE match_id > 0')

    data_files = [
        'Austerlitz_seizoen_2016-2017.xlsx',
        'Austerlitz_seizoen_2017-2018.xlsx',
        'Austerlitz_seizoen_2018-2019.xlsx',
    ]

    data_dir = os.path.join(os.path.dirname(__file__), '../data')

    for file_name in data_files:
        # pprint(file_name)
        wb = openpyxl.load_workbook(
            "%s/%s" % (data_dir, file_name), read_only=True)

        # every sheet is a competition, add a new competition, name is the
        # filename
        comp = dartsense.competition.Competition(name=file_name)
        comp.save()

        for ws in wb.worksheets:
            rowdata = []
            header = [cell.value for cell in ws[1]]
            # print(header)

            if 'Speler1' in header and 'Speler2' in header:
                # LeagueRound
                event = dartsense.event.LeagueRound(
                    name=ws.title,
                    competition=comp
                )
            elif 'Speler' in header:
                # LeagueAdjust
                event = dartsense.event.LeagueAdjust(
                    name=ws.title,
                    competition=comp
                )

            else:
                continue

            event.save()

            title_list = ws.title.split()
            date_avond = dateutil.parser.parse(title_list[-1])

            iter_rows = ws.rows
            next(iter_rows)

            for row in iter_rows:
                # values = []
                # for col in range(1,len(colnames)+1):
                #     cell = ws.cell(column=col, row=row)
                #     values.append(cell.value if cell.value else '')

                rowdict = dict(zip(header, map(lambda x: x.value, row)))

                if event.type == 'league_round':

                    # print(rowdict)

                    if 'Speler' in rowdict and not rowdict['Speler']:
                        continue
                    if 'Speler1' in rowdict and not rowdict['Speler1']:
                        continue
                    if 'Speler2' in rowdict and not rowdict['Speler2']:
                        continue

                    speler1 = find_player(rowdict['Speler1'], comp.id)
                    speler2 = find_player(rowdict['Speler2'], comp.id)

                    # pprint(speler1.name)
                    # pprint(speler2.name)

                    if speler1 == None:
                        print(rowdict['Speler1'])

                    if speler2 == None:
                        print(rowdict['Speler2'])

                    match = dartsense.match.Match()
                    match.event = event
                    match.player_1 = speler1
                    match.player_1_score = rowdict['Legs1']
                    match.player_1_180s = rowdict['Max1']
                    match.player_1_lollies = rowdict['Lollies1']
                    match.player_1_finishes = str(rowdict['Finishes1']).split(
                        ',') if rowdict['Finishes1'] else []

                    match.player_2 = speler2
                    match.player_2_score = rowdict['Legs2']
                    match.player_2_180s = rowdict['Max2']
                    match.player_2_lollies = rowdict['Lollies2']
                    match.player_2_finishes = str(rowdict['Finishes2']).split(
                        ',') if rowdict['Finishes2'] else []

                    match.date = date_avond
                    match.save()

                elif event.type == 'league_adjust':
                    # todo: leage adjust nav toernooi oid
                    pass


def usage():
    print('''read_file.py 
        --file <filename>
        ''')
    sys.exit()

if __name__ == "__main__":
    main(sys.argv[1:])
