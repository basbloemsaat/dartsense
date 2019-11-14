#!/usr/bin/env python

import getopt
import inspect
import json
import ntpath
import os
import re
import sys
import time

from pprint import pprint

import openpyxl
import dateutil.parser

sys.path.append(os.path.join(os.path.dirname(__file__), "../lib"))


# print(os.path.dirname(__file__))

import dartsense


def main(argv):
    parameters = {
        'filename': '',
    }

    try:
        opts, args = getopt.getopt(
            argv, "hsp:v:",
            [
                "help", "file="])
    except getopt.GetoptError as err:
        print(err)
        usage()

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            usage()
        elif opt in ("-f", "--file"):
            parameters['filename'] = arg

    # pprint(parameters)

    if not parameters['filename']:
        usage()

    sheet_data = []

    wb = openpyxl.load_workbook(parameters['filename'], read_only=True)
    # pprint(wb)

    for sheet in wb.worksheets:
        # pprint(sheet)
        # print(sheet.title)
        title_list = sheet.title.split()
        pprint(title_list)

        date_avond = dateutil.parser.parse(title_list[-1])
        if len(title_list) > 1:
            type_avond = title_list[0].lower()
        else:
            type_avond = "regulier"

        header = [cell.value for cell in sheet[1]]

        # pprint(header)
        iter_rows = sheet.rows
        next(iter_rows)
        for row in iter_rows:
            values = {
                "Date": date_avond.strftime('%Y-%m-%d'),
                "Type": type_avond
            }
            for key, cell in zip(header, row):
                values[key] = cell.value 
                if values[key] == None:
                    values[key] = 0
            
            sheet_data.append(values)

    # pprint(sheet_data)

    if(sheet_data):
        filename = re.sub(r'xlsx$', r'json', parameters['filename'])
        if (filename != parameters['filename']):
            filename = 'var/www/' + filename
            pprint(filename)
            with open(filename, 'w') as outfile:
                # pass
                json.dump(sheet_data, outfile)


def usage():
    print('''read_file.py 
        --file <filename>
        ''')
    sys.exit()

if __name__ == "__main__":
    main(sys.argv[1:])
