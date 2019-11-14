#!/usr/bin/env python

import getopt
import inspect
import os
import sys
import time

from pprint import pprint

import openpyxl

sys.path.append(os.path.join(os.path.dirname(__file__), "../lib"))


print(os.path.dirname(__file__))

import dartsense

def main(argv):
    parameters = {
        'filename': '',
    }

    try:
        opts, args = getopt.getopt(
            argv, "hsp:v:",
            [
                "help", "file="])
    except getopt.GetoptError as err:
        print(err)
        usage()

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            usage()
        elif opt in ("-f", "--file"):
            parameters['filename'] = arg

    pprint(parameters)

    if not parameters['filename']:
        usage()        

    wb = openpyxl.load_workbook(parameters['filename'], read_only=True)
    # pprint(wb)

    for sheet in wb.worksheets:
        # pprint(sheet)
        print(sheet.title)

        header = [cell.value for cell in sheet[1]]

        # pprint(header)
        iter_rows = sheet.rows
        next(iter_rows)
        for row in iter_rows:
            values = {}
            for key, cell in zip(header, row):
                pass
                values[key] = cell.value
            pprint(values)



def usage():
    print('''read_file.py 
        --file <filename>
        ''')
    sys.exit()

if __name__ == "__main__":
    main(sys.argv[1:])
