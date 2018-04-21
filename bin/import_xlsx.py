#!/usr/bin/env python3

import getopt
import inspect
import json
import ntpath
import os
import re
import sys
import time

from pprint import pprint
from dumper import dump

import openpyxl
import dateutil.parser

sys.path.append(os.path.join(os.path.dirname(__file__), "../lib"))


print(os.path.dirname(__file__))

import dartsense
import dartsense.player
import dartsense.event
import dartsense.match

players = {}


def find_player(player_name, competition_id):
    speler = None
    if player_name in players:
        speler = players[player_name]
    else:
        player_list = dartsense.player.PlayerList(
            filters={'competition': competition_id},
            search=player_name
        )

        # dump(player_list)

        if len(player_list) == 1:
            speler = players[player_name] = player_list[0]
        elif(len(player_list) > 1):
            for player in player_list:
                # pprint('0',player.name)
                # pprint('1',player_name)
                if player.name == player_name:
                    speler = player
                    break
    return speler


def main(argv):
    parameters = {
        'filename': '',
        'competition_id': 0,
    }

    try:
        opts, args = getopt.getopt(
            argv, "hsp:v:",
            ["help", "file=", "competition="])
    except getopt.GetoptError as err:
        print(err)
        usage()

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            usage()
        elif opt in ("-f", "--file"):
            parameters['filename'] = arg
        elif opt in ("-c", "--competition"):
            print('comp')
            competition_id = int(arg)

    pprint(competition_id)

    if not parameters['filename'] or not competition_id:
        usage()

    wb = openpyxl.load_workbook(parameters['filename'], read_only=True)
    # pprint(wb)

    for sheet in wb.worksheets:
        # pprint(sheet)
        # print(sheet.title)
        title_list = sheet.title.split()

        date_avond = dateutil.parser.parse(title_list[-1])
        if len(title_list) > 1:
            type_avond = title_list[0].lower()
            event = dartsense.event.LeagueAdjust()
        else:
            type_avond = "regulier"
            event = dartsense.event.LeagueRound()

        event.name = date_avond.strftime('%Y-%m-%d')
        event.competition = competition_id
        event.save()
        # print(date_avond.strftime('%Y-%m-%d'))
        # print(type_avond)

        header = [cell.value for cell in sheet[1]]

        # pprint(header)
        iter_rows = sheet.rows
        next(iter_rows)
        for row in iter_rows:
            values = {}
            for key, cell in zip(header, row):
                # pass
                values[key] = cell.value
            pprint(values)
            speler1 = find_player(values['Speler1'], competition_id)
            speler2 = find_player(values['Speler2'], competition_id)

            # pprint(speler1.name)
            # pprint(speler2.name)

            if speler1 == None:
                print(values['Speler1'])

            if speler2 == None:
                print(values['Speler2'])

            match = dartsense.match.Match()
            match.event = event
            match.player_1 = speler1
            match.player_1_score = values['Legs1']
            match.player_1_180s = values['Max1']
            match.player_1_lollies = values['Lollies1']
            match.player_1_finishes = str(values['Finishes1']).split(',') if values['Finishes1'] else []

            match.player_2 = speler2
            match.player_2_score = values['Legs2']
            match.player_2_180s = values['Max2']
            match.player_2_lollies = values['Lollies2']
            match.player_2_finishes = str(values['Finishes2']).split(',') if values['Finishes2'] else []

            match.date = date_avond
            match.save()



def usage():
    print('''read_file.py 
        --file <filename>
        ''')
    sys.exit()

if __name__ == "__main__":
    main(sys.argv[1:])
