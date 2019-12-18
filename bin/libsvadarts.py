

# import inspect
import json
# import ntpath
import os
# import re
import sqlite3
# import time

import dateutil.parser
import openpyxl
from pprint import pprint

dartssqlitedb = os.path.join(
    os.path.dirname(__file__),
    "../data/svadarts.sqlite3db"
)

# uri = 'file:' + dartssqlitedb + '?mode=ro'
# db = sqlite3.connect(uri, uri=True)
db = sqlite3.connect(dartssqlitedb)
# db.isolation_level = None

aliases = {
    "Andor": "Andor", "Anil": "Anil", "Bas": "Bas", "Bert": "Bert", "Brandon": "Brandon", "Christa": "Christa", "Colin": "Colin", "Ed": "Ed", "Elbert": "Elbert", "Erik H": "Erik H", "Erik": "Erik H", "Ernie": "Ernie", "Frank": "Frank", "Frans": "Frans", "Gert J": "Gert J", "Gert": "Gert J", "Gijs": "Gijs", "Gilbert": "Gilbert", "Hans": "Hans", "Harry": "Harry", "Henri": "Henri", "Jari": "Jari", "Johan": "Johan", "John": "John", "Joost": "Joost", "Joris": "Joris", "JR": "JR", "Kim": "Kim", "Maik": "Maik", "Marianne": "Marianne", "Martin": "Martin", "Menno": "Menno", "Otto": "Otto", "Paplip": "Paplip", "Pietra": "Pietra", "Reyn": "Reyn", "Youri": "Youri"}


def load_xlsx(filename):
    wb = openpyxl.load_workbook(filename, read_only=True)
    sheet_data = []
    for sheet in wb.worksheets:
        title_list = sheet.title.split()

        date_avond = dateutil.parser.parse(title_list[-1])
        if len(title_list) > 1:
            type_avond = title_list[0].lower()
        else:
            type_avond = "regulier"

        header = [cell.value for cell in sheet[1]]

        # pprint(header)
        iter_rows = sheet.rows
        next(iter_rows)
        for row in iter_rows:
            values = {
                "Date": date_avond.strftime('%Y-%m-%d'),
                "Type": type_avond
            }
            for key, cell in zip(header, row):
                values[key] = cell.value
                if values[key] == None:
                    values[key] = 0

            sheet_data.append(values)

    return sheet_data


def save_data_to_json(data, filename):
    # pprint(filename)
    with open(filename, 'w') as outfile:
        json.dump(data, outfile)


def exec_select_query(query, args=[], json_file=None):
    c = db.cursor()
    c.execute(query, args)
    names = [d[0] for d in c.description]
    rows = [dict(zip(names, row)) for row in c.fetchall()]

    if json_file:
        save_to_json(rows, json_file)
    return rows


def init_clean_db():
    # os.remove(dartssqlitedb)
    # db = sqlite3.connect(dartssqlitedb)

    db.execute('''
        DROP TABLE IF EXISTS speler
    ''')   
    db.execute('''
        DROP TABLE IF EXISTS game
    ''') 
    db.execute('''
        DROP TABLE IF EXISTS adjustments
    ''')

    db.execute('''
        CREATE TABLE speler (
            speler_naam VARCHAR(128)
        ) ''')

    db.execute('''
        CREATE TABLE game (
            game_id         INTEGER PRIMARY KEY,
            game_order      INT,
            datum           DATE,
            round           VARCHAR(16),
            speler1_naam    VARCHAR(128),
            speler2_naam    VARCHAR(128),
            speler1_legs    INT,
            speler2_legs    INT,
            speler1_180s    INT,
            speler2_180s    INT,
            speler2_lollies INT,
            speler1_lollies INT,
            speler2_finishes VARCHAR(32),
            speler1_finishes VARCHAR(32),
            speler1_punten  INT,
            speler2_punten  INT,
            speler1_rating  INT,
            speler2_rating  INT,
            speler1_rating_adj  INT,
            speler2_rating_adj  INT,

            UNIQUE (datum, round, speler1_naam, speler2_naam)
        ) ''')
    db.execute('''
        CREATE TABLE adjustments (
            adj_id          INTEGER PRIMARY KEY,
            datum           DATE,
            adj_type        VARCHAR(32),
            speler_naam     VARCHAR(128),
            speler_points   INT,
            speler_180s     INT,
            speler_lollies  INT,
            speler_finishes VARCHAR(32),

            UNIQUE (datum, speler_naam)
        ) ''')

    db.commit()

def load_all_data_into_db():
    # db = sqlite3.connect(dartssqlitedb)
    files = ['Austerlitz_seizoen_2016-2017.xlsx', 'Austerlitz_seizoen_2017-2018.xlsx',
             'Austerlitz_seizoen_2018-2019.xlsx', 'Austerlitz_seizoen_2019-2020.xlsx']

    spelers = {

    }

    def get_speler_naam(naam):
        alias = aliases[naam]
        spelers[alias] = 1
        return aliases[naam]

    for file in files:
        xlsx_path = os.path.join(
            os.path.dirname(__file__),
            "../data/" + file
        )

        data = load_xlsx(xlsx_path)
        order = 1;
        for entry in data:
            if entry['Type'] == 'regulier':
                if entry['Speler1'] == 0:
                    # soms meer regels in sheet: overslaan
                    continue

                db.execute('''INSERT INTO game (
                    game_order, datum, round, speler1_naam, speler2_naam,
                    speler1_legs, speler2_legs, speler1_180s, speler2_180s, speler2_lollies,
                    speler1_lollies, speler2_finishes, speler1_finishes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', [
                    order, 
                    entry['Date'],
                    entry['Ronde'],
                    get_speler_naam(entry['Speler1']),
                    get_speler_naam(entry['Speler2']),
                    entry['Legs1'],
                    entry['Legs2'],
                    entry['Max1'],
                    entry['Max2'],
                    entry['Lollies1'],
                    entry['Lollies2'],
                    entry['Max1'],
                    entry['Max2']
                ])
                order = order + 1

            else:
                # pprint(entry)
                db.execute('''INSERT INTO adjustments (
                    datum, adj_type, speler_naam, speler_points, speler_180s, speler_lollies, speler_finishes) 
                    VALUES (?,?,?,?,?,?,?)
                ''', [
                    entry['Date'],
                    entry['Type'],
                    get_speler_naam(entry['Speler']),
                    entry['Points'],
                    entry['Max'],
                    entry['Lollies'],
                    entry['Finishes'],
                ])

            # speler
        db.commit()

    for speler in spelers:
        db.execute('''INSERT INTO speler (speler_naam) VALUES (?)''', [speler])
        # pprint(speler)

    db.commit();

    # pprint(order)