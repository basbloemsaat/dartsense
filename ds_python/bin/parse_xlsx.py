#!/usr/bin/env python3

import getopt
import inspect
import json
import ntpath
import os
import re
import sys
import time

from pprint import pprint

import openpyxl
import dateutil.parser

sys.path.append(os.path.join(os.path.dirname(__file__), "../lib"))


# print(os.path.dirname(__file__))

import dartsense


def main(argv):
    parameters = {
        'filename': '',
    }

    try:
        opts, args = getopt.getopt(
            argv, "hsp:v:",
            [
                "help", "file="])
    except getopt.GetoptError as err:
        print(err)
        usage()

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            usage()
        elif opt in ("-f", "--file"):
            parameters['filename'] = arg

    # pprint(parameters)

    if not parameters['filename']:
        usage()

    sheet_data = []

    wb = openpyxl.load_workbook(parameters['filename'], read_only=True)
    # pprint(wb)

    for sheet in wb.worksheets:
        # pprint(sheet)
        # print(sheet.title)
        date_avond = dateutil.parser.parse(sheet.title)
        # pprint(date_avond)

        header = [cell.value for cell in sheet[1]]

        # pprint(header)
        iter_rows = sheet.rows
        next(iter_rows)
        for row in iter_rows:
            values = {
                "Date": date_avond.strftime('%Y-%m  -%d')
            }
            for key, cell in zip(header, row):
                values[key] = cell.value
                sheet_data.append(values)

    if(sheet_data):
        filename = re.sub(r'xlsx$',r'json', parameters['filename'])
        if (filename != parameters['filename']):
            filename = 'var/www/' + filename
            pprint(filename)
            with open(filename, 'w') as outfile:
                # pass
                json.dump(sheet_data, outfile)

def usage():
    print('''read_file.py 
        --file <filename>
        ''')
    sys.exit()

if __name__ == "__main__":
    main(sys.argv[1:])
